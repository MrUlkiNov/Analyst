import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def create_executive_report():
    """ Создает Excel отчет """
    results_df = pd.read_excel('prolongation_analysis_results.xlsx', sheet_name='Детальные_данные')
    yearly_avg = pd.read_excel('prolongation_analysis_results.xlsx', sheet_name='Средние_за_год')
    manager_stats = pd.read_excel('prolongation_analysis_results.xlsx', sheet_name='Статистика_менеджеров')
    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, color="366092", size=16)
    metric_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    blue_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    green_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')
    ws_summary = wb.create_sheet("Сводка для руководителя")
    ws_summary.merge_cells('A1:F1')
    ws_summary['A1'] = "АНАЛИТИЧЕСКИЙ ОТЧЕТ ПО ПРОЛОНГАЦИЯМ"
    ws_summary['A1'].font = title_font
    ws_summary['A1'].alignment = center_align
    ws_summary.merge_cells('A2:F2')
    ws_summary['A2'] = "2023 год"
    ws_summary['A2'].font = Font(bold=True, size=14, color="366092")
    ws_summary['A2'].alignment = center_align
    ws_summary['A4'] = "КЛЮЧЕВЫЕ ПОКАЗАТЕЛИ ОТДЕЛА"
    ws_summary['A4'].font = metric_font
    total_k1 = yearly_avg[yearly_avg['AM'] == 'ВСЕ МЕНЕДЖЕРЫ']['Коэффициент 1 (1 месяц)'].iloc[0]
    total_k2 = yearly_avg[yearly_avg['AM'] == 'ВСЕ МЕНЕДЖЕРЫ']['Коэффициент 2 (2 месяц)'].iloc[0]
    metrics = [
        ["Период анализа", "Март - Декабрь 2023"],
        ["Количество менеджеров", f"{len(manager_stats)} чел."],
        ["Всего пролонгаций", f"{manager_stats['Количество пролонгаций'].sum()} шт."],
        ["Средний K1 по отделу", f"{total_k1:.1%}"],
        ["Средний K2 по отделу", f"{total_k2:.1%}"]
    ]
    for i, (metric, value) in enumerate(metrics, 5):
        ws_summary[f'A{i}'] = metric
        ws_summary[f'B{i}'] = value
        ws_summary[f'A{i}'].font = metric_font
        ws_summary[f'B{i}'].fill = blue_fill
    ws_summary['A11'] = "РЕЙТИНГ МЕНЕДЖЕРОВ ПО ЭФФЕКТИВНОСТИ"
    ws_summary['A11'].font = metric_font
    headers = ['Место', 'Менеджер', 'K1', 'K2', 'Пролонгаций', 'Оценка']
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(12, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    ranking = yearly_avg[yearly_avg['AM'] != 'ВСЕ МЕНЕДЖЕРЫ'].copy()
    ranking = ranking.merge(manager_stats, left_on='AM', right_on='Менеджер')
    ranking = ranking.sort_values('Коэффициент 1 (1 месяц)', ascending=False)
    for i, (_, row) in enumerate(ranking.iterrows(), 13):
        ws_summary.cell(i, 1, i - 12)  # Место
        ws_summary.cell(i, 2, row['AM'])  # Менеджер
        ws_summary.cell(i, 3, f"{row['Коэффициент 1 (1 месяц)']:.1%}")  # K1
        ws_summary.cell(i, 4, f"{row['Коэффициент 2 (2 месяц)']:.1%}")  # K2
        ws_summary.cell(i, 5, row['Количество пролонгаций'])  # Количество
        k1 = row['Коэффициент 1 (1 месяц)']
        if k1 > 0.3:
            assessment = "Высокая"
            fill = green_fill
        elif k1 > 0.1:
            assessment = "Средняя"
            fill = blue_fill
        else:
            assessment = "Низкая"
            fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
        ws_summary.cell(i, 6, assessment).fill = fill
        for col in range(1, 7):
            cell = ws_summary.cell(i, col)
            if col in [3, 4]:
                cell.fill = blue_fill

    """ Лист 2 с динамикой """
    ws_trends = wb.create_sheet("Динамика")

    dept_data = results_df[results_df['AM'] == 'ВСЕ МЕНЕДЖЕРЫ'].copy()
    dept_data = dept_data.sort_values('Period')
    ws_trends.merge_cells('A1:E1')
    ws_trends['A1'] = "ДИНАМИКА КОЭФФИЦИЕНТОВ ПО МЕСЯЦАМ"
    ws_trends['A1'].font = title_font
    ws_trends['A1'].alignment = center_align
    headers = ['Месяц', 'K1', 'K2', 'Тренд K1', 'Тренд K2']
    for col, header in enumerate(headers, 1):
        cell = ws_trends.cell(3, col, header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    for i, (_, row) in enumerate(dept_data.iterrows(), 4):
        ws_trends.cell(i, 1, row['Period'])
        ws_trends.cell(i, 2, f"{row['Коэффициент 1 (1 месяц)']:.1%}").fill = blue_fill
        ws_trends.cell(i, 3, f"{row['Коэффициент 2 (2 месяц)']:.1%}").fill = blue_fill
        k1 = row['Коэффициент 1 (1 месяц)']
        k2 = row['Коэффициент 2 (2 месяц)']
        trend_k1 = "↑" if k1 > 0.2 else "↓" if k1 < 0.1 else "→"
        trend_k2 = "↑" if k2 > 0.1 else "↓" if k2 < 0.05 else "→"
        ws_trends.cell(i, 4, trend_k1)
        ws_trends.cell(i, 5, trend_k2)
    column_widths = {
        'A': 15, 'B': 20, 'C': 10, 'D': 10, 'E': 10, 'F': 15
    }
    for ws in wb.worksheets:
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
    wb.save('Отчет_по_пролонгациям_для_руководителя.xlsx')
    print("Отчет для руководителя создан!")

if __name__ == "__main__":
    create_executive_report()